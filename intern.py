from openpyxl import Workbook
import openpyxl
from definitions import flavors_dict, angles_dict, CERNA
from fwk_colors import summary, OOI
import streamlit as st
import OAI
from pymongo import MongoClient

flavors = list(flavors_dict)

angles = {
    "Level of Analysis": ["Micro", "Meso", "Macro"],
    "Spatial scope": ["Local", "Regional", "National", "International"],
    "Time Horizon": ["Short term", "Mid term", "Long term"],
}

h = OAI.Helper("wp2_assistant")

h.GOTOCACHE = st.secrets["CACHE"]
h.DBAdress = st.secrets["DB"]
h.DB = st.secrets["DB"]
h.PWD = st.secrets["PWD"]
h.cluster = MongoClient(h.DBAdress)
h.db = h.cluster["OAI"]
h.collection = h.db["OAI_Collection"]
h.DB = h.collection


def createBackground(flavor, angle):
    background = (
        """# Background

## Your role

You are an international expert in behavioral science and social innovation.
You are part of a team of experts working on the assessment of a social innovation, which is detailed below. You role is to provide a concise but very accurate assessment of the text.
You will given a "flavor" (as part of the PESTLE framework), as well as an "angle" to do this review.
For now, the "flavor" will be to do your assessment considering the '"""
        + flavor
        + """' flavor, and you will want to consider the """
        + angle
        + """ angle."""
    )
    background += "\n\n## Definitions\n\n"

    background += (
        "* '" + flavor + "' can capture topics like: '" + flavors_dict[flavor] + "'.\n"
    )
    background += "* '" + angle + "' can capture topics at different scales like:\n"
    for specificangle in angles[angle]:
        background += (
            "  * '"
            + specificangle
            + "': captures '"
            + angles_dict[angle][specificangle]
            + "'\n"
        )
    # print(background)
    background2 = (
        """

# Expected answer

You will need to start your answer as a single line, starting with a '*' bulletpoint, containing only one of the following items: '"""
        + ", ".join(angles[angle])
        + """'. This item should be the scale that is the most relevant for the """
        + angle
        + """ angle.
Then, add another bulletpoint, and write a short (up to 3 sentences) detailing as much as you can the content of the social innovation considering the '"""
        + flavor
        + """' flavor and the '"""
        + angle
        + """' angle. It must not start with "The text contains" or "This text contains" but rather directly go to the conclusion.
If there is nothing really relevant, just say so. Do not invent anything, and explain what is missing and what you'd like to see.

# Example of answer

## Example 1
* Mid term
* Expected economic results of the innovation are indicated between 3 to 7 years.

# Now is your time to work! The text is as below:

"""
    )

    return background + background2


def augmentReview(CERNA_review):
    F, A = flavors, list(angles.keys())
    for k in range(len(F)):
        flavorsummary = "* " + F[k] + "\n"
        for j in range(len(A)):
            flavorsummary += (
                "  * "
                + A[j]
                + " : "
                + CERNA_review[F[k]][A[j]].split("\n")[-1].strip("*").strip()
                + "\n"
            )
        CERNA_review[F[k]]["exec_summary"] = h.ask(
            "# Instructions:\n\n"
            + summary[0]
            + "\n\nIt must be particularly relevant to the '"
            + F[k]
            + "' topic and focus on it.\n\n# Content on which instructions apply:",
            flavorsummary,
            v="gpt-3.5-turbo-16k-0613",
            ow=False,
            src="none",
            seed="",
        )
        CERNA_review[F[k]]["output"] = h.ask(
            "# Instructions:\n\n"
            + OOI[0]
            + "\n\n# Content on which instructions apply:",
            flavorsummary,
            v="gpt-3.5-turbo-16k-0613",
            ow=False,
            src="none",
            seed="",
        )
        CERNA_review[F[k]]["output_score"] = h.ask(
            "# Instructions:\n\nRead the following content, and answer 1 if it says 'not satisfactory or irrelevant', 2 if it says 'satisfatory but improvements' and 3 if it says 'highly satisfactory'\n\n# Content on which instructions apply:",
            CERNA_review[F[k]]["output"],
            v="gpt-3.5-turbo-16k-0613",
            ow=False,
            src="none",
            seed="",
        )

        CERNA_review[F[k]]["outcome"] = h.ask(
            "# Instructions:\n\n"
            + OOI[1]
            + "\n\n# Content on which instructions apply:",
            flavorsummary,
            v="gpt-3.5-turbo-16k-0613",
            ow=False,
            src="none",
            seed="",
        )
        CERNA_review[F[k]]["outcome_score"] = h.ask(
            "# Instructions:\n\nRead the following content, and answer 1 if it says 'not satisfactory or irrelevant', 2 if it says 'satisfatory but improvements' and 3 if it says 'highly satisfactory'\n\n# Content on which instructions apply:",
            CERNA_review[F[k]]["outcome"],
            v="gpt-3.5-turbo-16k-0613",
            ow=False,
            src="none",
            seed="",
        )
        CERNA_review[F[k]]["impact"] = h.ask(
            "# Instructions:\n\n"
            + OOI[2]
            + "\n\n# Content on which instructions apply:",
            flavorsummary,
            v="gpt-3.5-turbo-16k-0613",
            ow=False,
            src="none",
            seed="",
        )
        CERNA_review[F[k]]["impact_score"] = h.ask(
            "# Instructions:\n\nRead the following content, and answer 1 if it says 'not satisfactory or irrelevant', 2 if it says 'satisfatory but improvements' and 3 if it says 'highly satisfactory'\n\n# Content on which instructions apply:",
            CERNA_review[F[k]]["impact"],
            v="gpt-3.5-turbo-16k-0613",
            ow=False,
            src="none",
            seed="",
        )
        if len(CERNA_review[F[k]]["impact_score"]) > 3:
            CERNA_review[F[k]]["impact_score"] = CERNA_review[F[k]]["impact_score"][0:1]
        if len(CERNA_review[F[k]]["outcome_score"]) > 3:
            CERNA_review[F[k]]["outcome_score"] = CERNA_review[F[k]]["outcome_score"][
                0:1
            ]
        if len(CERNA_review[F[k]]["output_score"]) > 3:
            CERNA_review[F[k]]["output_score"] = CERNA_review[F[k]]["output_score"][0:1]
    return CERNA_review


def getWorkbook(CERNA_review, txt, Q):
    book = openpyxl.load_workbook("template.xlsx")
    F, A = flavors, list(angles.keys())
    ws = book["review"]

    chars = "ABCDEFGH"
    for k in range(len(F)):
        for j in range(len(A)):
            print(F[k], A[j], chars[k], j)
            ws[chars[k + 1] + str(2 * j + 3)] = (
                CERNA_review[F[k]][A[j]].split("\n")[-1].strip("*").strip()
            )
            ws[chars[k + 1] + str(2 * j + 2)] = (
                CERNA_review[F[k]][A[j]]
                .split("\n")[0]
                .strip("*")
                .strip()
                .replace('"', "")
            )
        ws[chars[k + 1] + "8"] = CERNA_review[F[k]]["exec_summary"]
        ws[chars[k + 1] + "9"] = CERNA_review[F[k]]["output_score"]
        ws[chars[k + 1] + "10"] = CERNA_review[F[k]]["output"]
        ws[chars[k + 1] + "11"] = CERNA_review[F[k]]["outcome_score"]
        ws[chars[k + 1] + "12"] = CERNA_review[F[k]]["outcome"]
        ws[chars[k + 1] + "13"] = CERNA_review[F[k]]["impact_score"]
        ws[chars[k + 1] + "14"] = CERNA_review[F[k]]["impact"]
    ws = book["text"]
    ws["A1"] = txt
    ws = book["QA"]
    QUESTIONS = list(Q.keys())
    for k in range(len(QUESTIONS)):
        q = QUESTIONS[k]
        ws["A" + str(k + 2)] = q
        ws["B" + str(k + 2)] = Q[q]
    return book
